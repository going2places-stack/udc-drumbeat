#!/usr/bin/env python3
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl

CALENDAR = Path("/Users/clawagent/Downloads/UnofficialDC_Love_Content_Calendar.xlsx")

FINISHED_IMAGES = {
    "03-24": {"title": "UDC_Quote_03-24_MLK.png", "id": "1G-s_L5sJv1V9kkuvBtgKEFdAQDmclnQZ"},
    "03-25": {"title": "UDC_Quote_03-25_LaoTzu.png", "id": "1EQ2IbuP4cGrDp8AQcQ0c3o0d2-1NEcvC"},
    "03-27": {"title": "UDC_Quote_03-27_Aristotle.png", "id": "1iJHpW2C1u0CnyA2okM0VIT1-P3xceUIb"},
    "03-29": {"title": "UDC_Quote_03-29_Rumi.png", "id": "1R_wdMtIQpgR0-BVFzSGksywnNbWSjOti"},
    "03-30": {"title": "UDC_Quote_03-30_Gandhi.png", "id": "1JCPx3el0sLmDzMMcDw8iOvXzlmYo6EnK"},
}


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.datetime.strptime(str(value), "%m/%d/%Y").date()


def main():
    wb = openpyxl.load_workbook(CALENDAR, data_only=True)
    ws = wb["Content Calendar"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values[0]:
            continue
        item = dict(zip(headers, values))
        source_date = parse_date(item["Date"])
        key = source_date.strftime("%m-%d")
        image = FINISHED_IMAGES.get(key)
        rows.append({"source_date": source_date.isoformat(), "key": key, "row": item, "image": image})

    start = dt.date.today() + dt.timedelta(days=1)
    ready = []
    missing = []
    for index, row in enumerate([row for row in rows if row["image"]]):
        item = row["row"]
        image = row["image"]
        ready.append(
            {
                "source_date": row["source_date"],
                "scheduled_date": (start + dt.timedelta(days=index)).isoformat(),
                "name": f"{row['key']} {item['Attribution']}",
                "content_type": item["Content Type"],
                "caption": item["Caption"] or "",
                "hashtags": item["Hashtag Set"] or "",
                "image_title": image["title"],
                "image_file_id": image["id"],
                "image_url": f"https://drive.google.com/uc?export=download&id={image['id']}",
            }
        )
    for row in rows:
        if not row["image"]:
            item = row["row"]
            missing.append(
                {
                    "source_date": row["source_date"],
                    "content_type": item["Content Type"],
                    "content": item["Content/Quote"],
                    "attribution": item["Attribution"],
                }
            )

    Path("data").mkdir(exist_ok=True)
    Path("docs").mkdir(exist_ok=True)
    Path("data/ready_posts.json").write_text(json.dumps(ready, indent=2) + "\n")
    Path("data/missing_visuals.json").write_text(json.dumps(missing, indent=2) + "\n")

    lines = [
        "# Calendar Inventory",
        "",
        f"Calendar rows read: {len(rows)}",
        f"Ready rows with matched finished images: {len(ready)}",
        f"Rows left unscheduled because no matching finished image was found: {len(missing)}",
        "",
        "## Matched Rows",
        "",
        "| Source Date | Scheduled Date | Content Type | Image Found | File ID |",
        "| --- | --- | --- | --- | --- |",
    ]
    for row in ready:
        lines.append(
            f"| {row['source_date']} | {row['scheduled_date']} | {row['content_type']} | {row['image_title']} | {row['image_file_id']} |"
        )
    lines.extend(["", "## Missing Visual Rows", ""])
    by_type = {}
    for row in missing:
        by_type[row["content_type"]] = by_type.get(row["content_type"], 0) + 1
    for content_type, count in sorted(by_type.items()):
        lines.append(f"- {content_type}: {count}")
    lines.append("")
    Path("docs/calendar_inventory.md").write_text("\n".join(lines))


if __name__ == "__main__":
    main()
